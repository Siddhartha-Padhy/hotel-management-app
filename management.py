# Import libraries
from kivy.lang.builder import Builder   #for building the main app
from kivymd.app import MDApp        #for kivymd
from kivy.uix.screenmanager import Screen, ScreenManager        #for multiple screens
from kivymd.uix.list import ThreeLineListItem       #for List in Records section
from kivy.uix.label import Label        #for adding Label widgets
from kivy.uix.popup import Popup        #for pop-ups
from datetime import date, datetime     #for storing the dates and using the current date
import openpyxl as xl       #for operating the excel file through the app
from kivymd.uix.snackbar import Snackbar


'''
wb_address: stores the address of the Excel workbook
ws_customer: stores the address of the sheet containing customer data
ws_employee: stores the address of the sheet containing employee data
'''
wb_address ="C:/Users/Admin/Desktop/record_book.xlsx"
ws_customer = "Sheet1"
ws_employee = "Sheet2"


''' The Login Screen for the employees '''
class VerifyScreen(Screen):
    def invalid_popup(self):      
        popup = Popup(title ="Invalid Credentials",content=Label(text="Please enter correct details"), size_hint=(0.5,0.3))  
        popup.open()
        self.clear()

    # Verify the user's login details by matching it to employee records in ws_employee sheet.
    # If wrong show popup else take to InnerScreen.
    def login(self):
        user = self.manager.get_screen('VerifyScreen').ids.user.text
        password = self.manager.get_screen('VerifyScreen').ids.password.text
        wb = xl.load_workbook(wb_address)
        ws = wb[ws_employee]
        valid = False
        for i in range(1,ws.max_row+1):
            if(ws[f'A{i}'].value==user and str(ws[f'B{i}'].value)==password):
                self.manager.current = 'InnerScreen'
                self.manager.transition.direction = 'left'
                global user_g
                user_g = user
                self.clear()
                valid = True
                break
        if valid==False:
            self.invalid_popup()

    #for clearing the TextField data
    def clear(self):
        self.manager.get_screen('VerifyScreen').ids.user.text=""
        self.manager.get_screen('VerifyScreen').ids.password.text=""


''' The Main Screen to be shown after a correct login '''
class InnerScreen(Screen):
    def invalid_popup(self):      
        popup = Popup(title ="Invalid Credentials",content=Label(text="Please enter correct details"), size_hint=(0.5,0.3))  
        popup.open()
        self.clear()

    def room_occupied(self):      
        popup = Popup(title ="Room Occupied",content=Label(text="Please enter some other room"), size_hint=(0.5,0.3))  
        popup.open()

    def clear(self):
        self.manager.get_screen('InnerScreen').ids.room_num.text=""
        self.manager.get_screen('InnerScreen').ids.amount.text=""

    def logout(self):
        self.manager.current = 'VerifyScreen'
        self.manager.transition.direction = 'right'

    '''
    * Display the records stored in excel file
    * First line contans room number, name of the guest and current bill
    * Second line contains email address and phone number
    * Third line contains guest's availability, number of members, check in date and check out date if they are available.
    '''
    def show_records(self):
        self.manager.get_screen('InnerScreen').ids.container.clear_widgets()
        wb = xl.load_workbook(wb_address)
        ws = wb[ws_customer]
        for i in range(2,ws.max_row+1):
            primary = "Room: "+str(ws[f'E{i}'].value)+" | "+ str(ws[f'A{i}'].value) + " | $" + str(ws[f'I{i}'].value)

            secondary = str(ws[f'B{i}'].value)+" | " + "Phone No.: " + str(ws[f'C{i}'].value)

            if str(ws[f'J{i}'].value) == "Yes":
                tertiary = "Available: "+str(ws[f'J{i}'].value) + " | Members: " + str(ws[f'D{i}'].value) + " | " + "Check in: " + str(datetime.strptime(str(ws[f'F{i}'].value), "%Y-%m-%d").date())
            else:
                tertiary = "Available: "+str(ws[f'J{i}'].value) + " | Members: " + str(ws[f'D{i}'].value) + " | " + "Check in: " + str(datetime.strptime(str(ws[f'F{i}'].value), "%Y-%m-%d").date()) + " | Check out: " + str(datetime.strptime(str(ws[f'G{i}'].value), "%Y-%m-%d").date())

            self.manager.get_screen('InnerScreen').ids.container.add_widget(
                ThreeLineListItem(text= primary, secondary_text = secondary, tertiary_text = tertiary)
            )

    #Displays the name of the guest once Enter key is pressed after entering the room number.
    def get_details(self):
        wb = xl.load_workbook(wb_address)
        ws = wb[ws_customer]
        key = self.manager.get_screen('InnerScreen').ids.room_num.text
        for i in range(2,ws.max_row+1):
            if(str(ws[f'E{i}'].value)==key):
                self.manager.get_screen('InnerScreen').ids.name_guest.text = str(ws[f'A{i}'].value)


    def is_available(self):
        key = self.manager.get_screen('InnerScreen').ids.room.text
        found = False
        wb = xl.load_workbook(wb_address)
        ws = wb[ws_customer]
        for i in range(2,ws.max_row+1):
            if key == str(ws[f'E{i}'].value) and str(ws[f'J{i}'].value) == "Yes":
                found = True
                break
        if found:
            return True
        else:
            return False


    #Adding new guest details to records and clear the TextFields after adding.
    #If room is already occupied show popup
    def add_guest(self):
        if not self.is_available():
            wb = xl.load_workbook(wb_address)
            ws = wb[ws_customer]
            top = ws.max_row + 1
            ws[f'A{top}'] = self.manager.get_screen('InnerScreen').ids.guest_name.text
            ws[f'B{top}'] = self.manager.get_screen('InnerScreen').ids.guest_email.text
            ws[f'C{top}'] = self.manager.get_screen('InnerScreen').ids.guest_phone.text
            ws[f'D{top}'] = self.manager.get_screen('InnerScreen').ids.guest_member_number.text
            ws[f'E{top}'] = self.manager.get_screen('InnerScreen').ids.room.text
            ws[f'F{top}'] = str(datetime.strptime(str(date.today()), "%Y-%m-%d").date())
            ws[f'H{top}'] = user_g
            ws[f'I{top}'] = 0
            ws[f'J{top}'] = "Yes"
            wb.save(wb_address)
            self.manager.get_screen('InnerScreen').ids.guest_name.text = ""
            self.manager.get_screen('InnerScreen').ids.guest_email.text = ""
            self.manager.get_screen('InnerScreen').ids.guest_phone.text = ""
            self.manager.get_screen('InnerScreen').ids.guest_member_number.text = ""
            self.manager.get_screen('InnerScreen').ids.room.text = ""
            Snackbar(text="Guest Added", bg_color=(40/255,40/255,43/255,1),snackbar_x="200dp",snackbar_y="30dp",size_hint_x=0.5).open()
        else:
            self.room_occupied()

    '''
    * Add amount to current bill of a guest.
    * Display the name of guest for confirmation.
    * If room is not occupied show popup
    '''
    def add_amount(self):
        wb = xl.load_workbook(wb_address)
        ws = wb[ws_customer]
        amount = self.manager.get_screen('InnerScreen').ids.amount.text
        room = self.manager.get_screen('InnerScreen').ids.room_num.text
        found = False
        for r in range(2,ws.max_row+1):
            if(str(ws[f'E{r}'].value) == room and str(ws[f'J{r}'].value)=="Yes"):
                ws[f'I{r}'] = int(str(ws[f'I{r}'].value)) + int(amount)
                wb.save(wb_address)
                found = True
                break
        if(found==False):
            self.invalid_popup()
        self.clear()

    #Move to the final payment screen PaymentScreen
    def pay_total(self):
        self.manager.current = 'PaymentScreen'
        self.manager.transition.direction = 'right'

class ContentNavigationDrawer(Screen):
    pass


''' The Payment Screen to be displayed when user clicks on Pay Total Amount button in the Billing Section'''
class PaymentScreen(Screen):
    def invalid_popup(self):      
        popup = Popup(title ="Invalid Credentials",content=Label(text="Please enter correct details"), size_hint=(0.5,0.3))  
        popup.open()

    #Get the details of the guest once the Enter key is pressed after entering the room number.
    def get_details(self):
        wb = xl.load_workbook(wb_address)
        ws = wb[ws_customer]
        key = self.manager.get_screen('PaymentScreen').ids.room.text
        for i in range(2,ws.max_row+1):
            if(str(ws[f'E{i}'].value)==key):
                self.manager.get_screen('PaymentScreen').ids.name.text = str(ws[f'A{i}'].value)
                self.manager.get_screen('PaymentScreen').ids.amount.text = str(ws[f'I{i}'].value)
                self.manager.get_screen('PaymentScreen').ids.check_in.text = str(datetime.strptime(str(ws[f'F{i}'].value), "%Y-%m-%d").date())
                self.manager.get_screen('PaymentScreen').ids.check_out.text = str(date.today())
        
    '''
    * Make the availablity of the guest to "No" and set the current date to check out date
    * If the room number is not found or is not occupied show popup
    * Clear the TextFields after submission
    '''
    def checkout(self):
        wb = xl.load_workbook(wb_address)
        ws = wb[ws_customer]
        key = self.manager.get_screen('PaymentScreen').ids.room.text
        found = False
        for i in range(2,ws.max_row+1):
            if(str(ws[f'E{i}'].value)==key and str(ws[f'J{i}'].value) == "Yes"):
                ws[f'J{i}'] = "No"
                ws[f'G{i}'] = str(datetime.strptime(str(date.today()), "%Y-%m-%d").date())
                wb.save(wb_address)
                found = True
                break
        if (found == False):
            self.invalid_popup()
        self.manager.get_screen('PaymentScreen').ids.room.text = ""
        self.manager.get_screen('PaymentScreen').ids.name.text = ""
        self.manager.get_screen('PaymentScreen').ids.amount.text = ""
        self.manager.get_screen('PaymentScreen').ids.check_in.text = ""
        self.manager.get_screen('PaymentScreen').ids.check_out.text = ""

''' Screen Manager '''
class MainScreen(ScreenManager):
    pass

class MainApp(MDApp):
    def build(self):
        self.theme_cls.theme_style = "Dark"
        self.theme_cls.primary_palette = "Red"
        self.theme_cls.primary_hue = "700"
        return Builder.load_file("management.kv")

if __name__=="__main__":
    MainApp().run()